import logging
import sqlite3
from config import superuser
from google_drive import backup_googledrive
from database import Database
from aiogram.types import Message, FSInputFile
from aiogram.filters.command import Command
from aiogram import Bot, Router
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from collections import Counter


db = Database('db.db')
stat = Router()

@stat.message(Command('archive_database'))
async def create_archive(message: Message):
    if (int(message.from_user.id) == superuser and list(*db.get_lang(message.from_user.id))[0] == 'RU') or \
            (str(message.from_user.id) in [str(db.get_admin_id_stat()[x][0]) for x in range(len(db.get_admin_id_stat()))] and list(*db.get_lang(message.from_user.id))[0] == 'RU'):
        src = sqlite3.connect('db.db')
        dst = sqlite3.connect('core/backup.db')
        with dst:
            src.backup(dst, pages=1)
        dst.close()
        src.close()
        await message.answer(text = 'База данных скопирована!')
        try:
            backup_googledrive('core/backup.db')
            await message.answer(text='База данных отправлена в Google drive!')
        except:
            pass


@stat.message(Command('statistics'))
async def document(message: Message, bot: Bot):
    if (int(message.from_user.id) == superuser and list(*db.get_lang(message.from_user.id))[0] == 'RU') or \
            (str(message.from_user.id) in [str(db.get_admin_id_stat()[x][0]) for x in range(len(db.get_admin_id_stat()))] and list(*db.get_lang(message.from_user.id))[0] == 'RU'):
        messages = db.statistics()
        top_users = db.statistics_users()
        most_users = [user[0] for user in top_users]
        message_texts = [message[0] for message in messages]
        popular_queries = Counter(message_texts).most_common(10)
        popular_users = Counter(most_users).most_common(10)
        wb = Workbook()
        ws = wb.active
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20
        ws['A1'] = 'Category'
        ws['B1'] = 'Number of usages'
        for x in range(len(popular_queries)):
            category = popular_queries[x][0].replace('• ', '')
            ws.append([category,popular_queries[x][1]])
        chart = BarChart()
        labels = Reference(ws, min_col=1, min_row=2, max_row=11)
        data = Reference(ws, min_col=2, min_row=2, max_row=11)
        chart.add_data(data)
        chart.set_categories(labels)
        chart.title = " "
        ws.add_chart(chart, "D1")
        ws['A16'] = 'Most users'
        ws['B16'] = 'Number of usages'
        for x in range(len(popular_users)):
            try:
                user_first_name = list(*db.get_user_firstname(int(popular_users[x][0])))[0]
                user_last_name = list(*db.get_user_lastname(int(popular_users[x][0])))[0]
                lastname = user_last_name if user_last_name != None else ' '
                firstname = user_first_name if user_first_name != None else ' '
                user_name = f'{firstname} {lastname}'
                ws.append([user_name,popular_users[x][1]])
            except:
                pass
        chart = BarChart()
        labels = Reference(ws, min_col=1, min_row=17, max_row=26)
        data = Reference(ws, min_col=2, min_row=17, max_row=26)
        chart.add_data(data)
        chart.set_categories(labels)
        chart.title = " "
        ws.add_chart(chart, "D16")
        ws['A31'] = 'Number of users'
        ws['B31'] = f'{db.get_count_users()[0]}'
        wb.save("statistics.xlsx")
        doc = FSInputFile("statistics.xlsx")
        await bot.send_document(chat_id=message.from_user.id, document=doc)